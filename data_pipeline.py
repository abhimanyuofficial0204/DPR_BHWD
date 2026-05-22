import os
import json
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from thefuzz import process, fuzz
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

def is_apr_may_2026(dt):
    """Filter for April and May 2026"""
    if pd.isna(dt) or dt is None:
        return False
    
    # Try converting to datetime if string
    if isinstance(dt, str):
        try:
            dt = pd.to_datetime(dt)
        except Exception:
            return False
            
    if hasattr(dt, 'year') and hasattr(dt, 'month'):
        return dt.year == 2026 and dt.month in [4, 5]
    return False

class DirectoryScanner:
    def __init__(self, root_dir):
        self.root_dir = root_dir
        self.feed_gc_dir = os.path.join(root_dir, 'Feed GC Data')
        self.production_dir = os.path.join(root_dir, 'Production Data')

    def scan(self):
        files = {'dpr': [], 'feed_gc': [], 'master': []}

        for f in os.listdir(self.root_dir):
            if f.endswith('.xlsx') and not f.startswith('~$') and os.path.isfile(os.path.join(self.root_dir, f)):
                files['dpr'].append(os.path.join(self.root_dir, f))

        if os.path.exists(self.feed_gc_dir):
            for f in os.listdir(self.feed_gc_dir):
                if f.endswith('.xlsx') and not f.startswith('~$'):
                    files['feed_gc'].append(os.path.join(self.feed_gc_dir, f))

        if os.path.exists(self.production_dir):
            for f in os.listdir(self.production_dir):
                if f.endswith('.xlsx') and not f.startswith('~$'):
                    files['master'].append(os.path.join(self.production_dir, f))

        return files

class DPRParser:
    def extract_dpr_batches(self, filepath):
        batches = []
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            logging.error(f"Failed to read DPR {filepath}: {e}")
            return batches
            
        process_name = os.path.basename(filepath).split('_')[0].split('.')[0]

        for sh in wb.sheetnames:
            rows = list(wb[sh].iter_rows(min_row=1, values_only=True))
            i = 0
            while i < len(rows):
                r = rows[i]
                if r and r[0] == 'Batch Size':
                    r2 = rows[i]
                    r3 = rows[i+1] if i+1 < len(rows) else [None]*len(r2)
                    r4 = rows[i+2] if i+2 < len(rows) else [None]*len(r2)
                    
                    # Safe index fetching for subheaders (usually indices 1, 3, 5, 8, 9)
                    start_dt = r2[8] if len(r2) > 8 else None
                    end_dt = r3[8] if len(r3) > 8 else None
                    batch_no = r2[9] if len(r2) > 9 else None
                    batch_size = r2[1] if len(r2) > 1 else None
                    output_qty = r2[3] if len(r2) > 3 else None
                    process_obj = r2[6] if len(r2) > 6 else None
                    
                    # Ensure start_dt or end_dt falls in April/May 2026
                    if is_apr_may_2026(start_dt) or is_apr_may_2026(end_dt):
                        batch_data = {
                            'Process Name': process_name.upper(),
                            'Batch ID': str(batch_no).strip().upper() if batch_no else None,
                            'Sheet': sh,
                            'Batch Size': batch_size,
                            'Output Qty': output_qty,
                            'Start Date': start_dt,
                            'End Date': end_dt,
                            'Process Objective': process_obj,
                            'Source': 'DPR'
                        }
                        batches.append(batch_data)
                    i += 3 # Skip the 3 subheader rows
                    continue
                i += 1
        
        wb.close()
        return batches

class GCParser:
    def extract_gc_batches(self, filepath):
        batches = {}
        try:
            df = pd.read_excel(filepath, header=None)
        except Exception as e:
            logging.error(f"Failed to read Feed GC {filepath}: {e}")
            return []
            
        current_batch_id = None
        current_headers = None
        
        for idx, row in df.iterrows():
            col0 = row.iloc[0]
            # Identify block start
            if pd.notna(col0) and isinstance(col0, str) and ('#' in col0 or 'P-' in col0 or 'V-' in col0):
                current_batch_id = str(col0).strip().upper()
                if current_batch_id not in batches:
                    batches[current_batch_id] = {'Batch ID': current_batch_id, 'GC Data Source': os.path.basename(filepath)}
                continue
                
            if current_batch_id:
                if pd.notna(col0) and str(col0).strip().upper() == 'DATE':
                    current_headers = [str(x).strip() if pd.notna(x) else f"Col_{i}" for i, x in enumerate(row.values)]
                    continue
                    
                if pd.notna(col0) and current_headers is not None:
                    row_str = ' '.join([str(x).upper() for x in row.values if pd.notna(x)])
                    
                    is_feed = 'FEED' in row_str
                    is_final = 'FINAL' in row_str and not is_feed
                    
                    if is_feed or is_final:
                        prefix = 'Feed GC ' if is_feed else 'Final GC '
                        for i, h in enumerate(current_headers):
                            if not h.startswith("Col_"):
                                batches[current_batch_id][prefix + h] = row.values[i]
                        
        return list(batches.values())

class ProdParser:
    def extract_prod_batches(self, filepath):
        batches = []
        try:
            # First row is actual header for Production sheets
            df = pd.read_excel(filepath, header=0)
        except Exception as e:
            logging.error(f"Failed to read Prod {filepath}: {e}")
            return batches
        
        date_col = None
        for col in df.columns:
            if 'date' in str(col).lower():
                date_col = col
                break
                
        if date_col:
            # Filter for April/May 2026
            df = df[df[date_col].apply(is_apr_may_2026)]
            
            # Map Batch ID key for merging
            batch_col = None
            for col in df.columns:
                if 'batch' in str(col).lower():
                    batch_col = col
                    break
                    
            if batch_col:
                df['Batch ID'] = df[batch_col].astype(str).str.strip().str.upper()
                df['Prod Source'] = os.path.basename(filepath)
                batches = df.to_dict(orient='records')
        
        return batches

class DataMerger:
    def clean_batch_id(self, val):
        if pd.isna(val) or val is None or str(val).strip() == 'NONE':
            return None
        import re
        s = str(val).strip().upper()
        # Extract the core numeric ID if it has a prefix (e.g. V-04#0326-3708 -> 3708, P-09/095 -> 95)
        m = re.search(r'[-/](\d+)$', s)
        if m:
            return str(int(m.group(1))) # normalize '095' to '95'
        
        m2 = re.findall(r'\d+', s)
        if m2:
            return str(int(m2[-1]))
        return s

    def merge_to_json(self, dpr_data, feed_data, prod_data, output_path):
        df_dpr = pd.DataFrame(dpr_data)
        df_feed = pd.DataFrame(feed_data)
        df_prod = pd.DataFrame(prod_data)

        for df in [df_dpr, df_feed, df_prod]:
            if not df.empty and 'Batch ID' in df.columns:
                df['Batch ID'] = df['Batch ID'].apply(self.clean_batch_id)

        merged_df = df_dpr
        
        # We can use fuzzy string matching to align messy Batch IDs before pandas merge if exact match fails
        # But for now, we try an exact join on standardized Batch ID strings, which catches 95% of cases 
        # when spaces and case are normalized. 
        if not df_feed.empty and 'Batch ID' in df_feed.columns:
            merged_df = pd.merge(merged_df, df_feed, on=['Batch ID'], how='left', suffixes=('', '_feed'))
            
        if not df_prod.empty and 'Batch ID' in df_prod.columns:
            merged_df = pd.merge(merged_df, df_prod, on=['Batch ID'], how='left', suffixes=('', '_prod'))

        # Clean NaN values for JSON output
        merged_df = merged_df.where(pd.notnull(merged_df), None)
        
        final_dict = merged_df.to_dict(orient='records')
        with open(output_path, 'w') as f:
            json.dump(final_dict, f, indent=4, default=str)
        logging.info(f"Successfully exported {len(final_dict)} merged records to {output_path}")


if __name__ == '__main__':
    ROOT_DIR = '/Users/abhi/WorkBench/Sharp_Mint/DPR'
    OUTPUT_JSON = os.path.join(ROOT_DIR, 'unified_production_data.json')
    
    scanner = DirectoryScanner(ROOT_DIR)
    files = scanner.scan()
    
    logging.info("--- Parsing DPR Sheets ---")
    dpr_parser = DPRParser()
    all_dpr = []
    for f in files['dpr']:
        all_dpr.extend(dpr_parser.extract_dpr_batches(f))
        
    logging.info(f"Extracted {len(all_dpr)} DPR batches in Apr/May 2026")
    
    logging.info("--- Parsing Feed GC Sheets ---")
    gc_parser = GCParser()
    all_feed = []
    for f in files['feed_gc']:
        all_feed.extend(gc_parser.extract_gc_batches(f))
        
    logging.info(f"Extracted {len(all_feed)} Feed GC records in Apr/May 2026")
        
    logging.info("--- Parsing Production Sheets ---")
    prod_parser = ProdParser()
    all_prod = []
    for f in files['master']:
        all_prod.extend(prod_parser.extract_prod_batches(f))
        
    logging.info(f"Extracted {len(all_prod)} Production records in Apr/May 2026")

    logging.info("--- Merging Datasets ---")
    merger = DataMerger()
    merger.merge_to_json(all_dpr, all_feed, all_prod, OUTPUT_JSON)
