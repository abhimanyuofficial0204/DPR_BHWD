import io
import os
import re
import json
import math
import warnings
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
from google.oauth2 import service_account

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

app = FastAPI(title="Sharp Mint Parsing API", version="1.0")

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 1. Initialize Firestore Connection
db = None
CREDENTIAL_PATH = "firebase-adminsdk.json"

# Check in parent directory or current directory
cred_file = CREDENTIAL_PATH
if not os.path.exists(cred_file):
    parent_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), CREDENTIAL_PATH)
    if os.path.exists(parent_path):
        cred_file = parent_path

env_creds = os.environ.get("FIREBASE_SERVICE_ACCOUNT")

try:
    if env_creds:
        creds_dict = json.loads(env_creds)
        creds_obj = service_account.Credentials.from_service_account_info(creds_dict)
        db = firestore.Client(credentials=creds_obj, database="prod123")
        print("Connected to Firestore via Environment Variable.")
    elif os.path.exists(cred_file):
        creds_obj = service_account.Credentials.from_service_account_file(cred_file)
        db = firestore.Client(credentials=creds_obj, database="prod123")
        print(f"Connected to Firestore via key file: {cred_file}")
    else:
        print("Warning: Firestore credentials not found. Local offline mock mode active.")
except Exception as e:
    print(f"Firestore Initialization Error: {e}")


# Helper clean dict for Firestore NaNs
def clean_dict(d):
    clean = {}
    for k, v in d.items():
        if pd.isna(v) or v is None:
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        if isinstance(v, datetime):
            clean[k] = v.isoformat()
        else:
            clean[k] = v
    return clean

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace('%', ''))
    except:
        return default

def is_lm_header(col):
    if not isinstance(col, str):
        return False
    s = col.upper().strip()
    return 'L-MENTHOL' in s or 'L MENTHOL' in s or s == 'LM' or s.startswith('LM ') or s.endswith(' LM') or ' LM ' in s or 'LM(' in s or '(LM)' in s

def is_ma_header(col):
    if not isinstance(col, str):
        return False
    s = col.upper().strip()
    return 'MENTHYL ACETATE' in s or 'MENTHYL-ACETATE' in s or s == 'MA' or s == 'M.A.' or s == 'M.A' or s.startswith('MA ') or s.endswith(' MA') or ' MA ' in s or 'MA(' in s or '(MA)' in s

def is_hpt_header(col):
    if not isinstance(col, str):
        return False
    s = col.upper().strip()
    return 'HEPTANE' in s or s == 'HPT' or s == 'H.P.T.' or s == 'H.P.T' or s.startswith('HPT ') or s.endswith(' HPT') or ' HPT ' in s or 'HPT(' in s or '(HPT)' in s

# Batch ID clean regex matching: Plant ID + Core ID
def get_plant_and_core_id(batch_str, default_plant=None):
    if not batch_str or pd.isna(batch_str) or str(batch_str).strip().upper() == 'NONE':
        return None, None
    s = str(batch_str).strip().upper()
    plant_match = re.search(r'([VP]-\d+)', s)
    plant = plant_match.group(1) if plant_match else default_plant
    m = re.findall(r'\d+', s)
    core_id = None
    if m:
        core_id = str(int(m[-1]))
    return plant, core_id

def calculate_ma_conversion(output_lm_pct, output_ma_pct):
    if output_lm_pct == 0 and output_ma_pct == 0:
        return 0.0
    moles_ma = output_ma_pct / 198.30
    moles_lm = output_lm_pct / 156.27
    if (moles_ma + moles_lm) == 0:
        return 0.0
    return (moles_ma / (moles_ma + moles_lm)) * 100.0


@app.get("/api/health")
def health():
    return {"status": "ok", "database_connected": db is not None}


@app.post("/api/upload")
async def parse_and_sync_file(file: UploadFile = File(...)):
    """
    Endpoint to receive raw Excel sheet, auto-detect type,
    parse batches/drums, and upsert them to Firestore database.
    """
    if not db:
        raise HTTPException(status_code=500, detail="Cloud Database connection is offline.")

    contents = await file.read()
    file_io = io.BytesIO(contents)
    
    try:
        # Load sheets to inspect headers and classify type
        wb = openpyxl.load_workbook(file_io, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Load first row of first sheet to classify layout
        df_init = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_names[0], nrows=5, header=None)
        
        # Convert first few cells to check keywords
        flat_header = []
        for idx, row in df_init.iterrows():
            flat_header.extend([str(x).strip().upper() for x in row.values if pd.notna(x)])
            
        file_io.seek(0)
        
        # 1. Classification & Ingestion Dispatcher
        if any("BATCH SIZE" in str(x) for x in flat_header):
            # DPR Sheet Structure (Block layouts)
            batches_count = process_dpr_file(file_io, file.filename)
            return {"status": "success", "type": "DPR", "batches_upserted": batches_count}
            
        elif any("FILLED BY GC" in str(x) for x in flat_header) or any("LOCATION CODE" in str(x) for x in flat_header):
            # Flavour, Fragrance, or MN-MLQ sheets (Drum-wise GCs)
            drums_count = process_flavour_fragrance_file(file_io, file.filename)
            return {"status": "success", "type": "FLAVOUR_FRAGRANCE_DRUMS", "drums_upserted": drums_count}
            
        elif any("INPUT FEED" in str(x) for x in flat_header) and any("DRUM/ WT" in str(x) or "DR NO" in str(x) for x in flat_header):
            # Production Log (New Pro tabular weights)
            drums_count = process_production_logs(file_io, file.filename)
            return {"status": "success", "type": "PRODUCTION_DRUMS", "drums_upserted": drums_count}
            
        elif any("DATE" in str(x) for x in flat_header) and len(df_init.columns) > 30:
            # GC Log structure (Date header block, feed/final)
            batches_count = process_gc_file(file_io, file.filename)
            return {"status": "success", "type": "GC_ANALYSIS", "batches_updated": batches_count}
            
        else:
            raise HTTPException(status_code=400, detail="Unrecognised sheet structure or column headers.")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Parsing Error: {str(e)}")


def process_dpr_file(file_content, filename):
    wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
    batches_ref = db.collection('batches')
    drums_ref = db.collection('drums')
    batches_count = 0
    
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: continue
        
        header_row = rows[0]
        lm_cols = [idx for idx, col in enumerate(header_row) if is_lm_header(col)][:1]
        ma_cols = [idx for idx, col in enumerate(header_row) if is_ma_header(col)][:1]
        hpt_cols = [idx for idx, col in enumerate(header_row) if is_hpt_header(col)][:1]
        
        i = 0
        while i < len(rows):
            r = rows[i]
            if r and len(r) > 1 and (r[0] == 'Batch Size' or r[1] == 'Batch Size'):
                shift = 1 if r[1] == 'Batch Size' else 0
                r1 = rows[i]
                r2 = rows[i+1] if i+1 < len(rows) else [None]*150
                
                raw_batch_no = str(r1[9 + shift] or r2[9 + shift] or '').strip()
                plant, core_id = get_plant_and_core_id(raw_batch_no, default_plant=sh)
                
                if not core_id or not plant:
                    i += 1
                    continue
                    
                batch_key = f"{plant}_{core_id}"
                header_input_weight = safe_float(r1[1 + shift])
                
                header_output_weight = 0.0
                if sh != 'V-24':
                    for r_hdr in [r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                        if r_hdr:
                            for c_idx, cell_val in enumerate(r_hdr):
                                if str(cell_val).strip().lower() == 'output qty' and c_idx + 1 < len(r_hdr):
                                    header_output_weight = safe_float(r_hdr[c_idx + 1])
                                    break
                
                start_dt = None
                end_dt = None
                for r_hdr in [rows[i-1] if i>0 else [], r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                    if r_hdr:
                        for c_idx, cell_val in enumerate(r_hdr):
                            val_str = str(cell_val).strip().lower()
                            if 'start date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime): start_dt = candidate
                            elif 'end date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime): end_dt = candidate
                
                input_drums = []
                output_drums = []
                j = i + 1
                while j < len(rows) and j < i + 4:
                    dr_row = rows[j]
                    if not dr_row: break
                    c0 = str(dr_row[0] or '').strip().lower()
                    if 'batch size' in c0:
                        j += 1
                        continue
                    break
                
                while j < len(rows):
                    dr_row = rows[j]
                    if not dr_row or (len(dr_row) > 1 and (dr_row[0] == 'Batch Size' or dr_row[1] == 'Batch Size')):
                        break
                    if all(val is None or str(val).strip() == '' for val in dr_row):
                        break
                        
                    in_desc = str(dr_row[0 + shift] or '').strip().upper()
                    in_dr_no = str(dr_row[1 + shift] or '').strip()
                    in_weight = safe_float(dr_row[3 + shift] if 3+shift < len(dr_row) else 0)
                    if in_weight > 0 or in_desc != '':
                        input_drums.append({'desc': in_desc, 'drum_no': in_dr_no, 'weight': in_weight})
                        
                    out_weight_val = dr_row[18 + shift] if 18+shift < len(dr_row) else 0
                    out_desc_col = 26 + shift
                    out_desc = str(dr_row[out_desc_col] if out_desc_col < len(dr_row) else '').strip().upper()
                    if out_desc == '' and 11+shift < len(dr_row):
                        out_desc = str(dr_row[11+shift] or '').strip().upper()
                        
                    out_weight = safe_float(out_weight_val)
                    out_dr_no = ''
                    if 15+shift < len(dr_row) and dr_row[15+shift]:
                        out_dr_no = str(dr_row[15+shift]).strip()
                    elif 11+shift < len(dr_row) and dr_row[11+shift]:
                        out_dr_no = str(dr_row[11+shift]).strip()
                        
                    out_lm = get_first_nonzero(dr_row, lm_cols, shift)
                    out_ma = get_first_nonzero(dr_row, ma_cols, shift)
                    out_hpt = get_first_nonzero(dr_row, hpt_cols, shift)
                    
                    if out_weight > 0 or out_desc != '':
                        output_drums.append({
                            'desc': out_desc,
                            'drum_no': out_dr_no,
                            'weight': out_weight,
                            'lm_gc': out_lm,
                            'ma_gc': out_ma,
                            'hpt_gc': out_hpt
                        })
                    j += 1
                
                feed_types = set([d['desc'] for d in input_drums if d['desc']])
                output_types = set([d['desc'] for d in output_drums if d['desc'] and d['desc'] != 'WATER'])
                raw_process = str(r1[6 + shift] or '').strip().upper()
                
                process_type = "UNKNOWN"
                if 'DLM' in feed_types: process_type = 'ENZYME_RXN'
                elif 'ADH' in feed_types: process_type = 'SRP'
                elif 'HPT' in feed_types and 'NHPT' in output_types: process_type = 'WASHING'
                elif 'DMM' in feed_types: process_type = 'DISTILLATION'
                
                if process_type == "UNKNOWN":
                    if 'GLR' in raw_process or 'ENZYME' in raw_process: process_type = 'ENZYME_RXN'
                    elif 'SRP' in raw_process: process_type = 'SRP'
                    elif 'WASH' in raw_process: process_type = 'WASHING'
                
                sum_in_drums = sum(d['weight'] for d in input_drums)
                sum_out_drums = sum(d['weight'] for d in output_drums)
                
                final_input = sum_in_drums if sum_in_drums > 0 else header_input_weight
                final_output = sum_out_drums if sum_out_drums > 0 else header_output_weight
                if process_type == 'ENZYME_RXN' and final_output == 0:
                    final_output = final_input
                
                total_out_weight = sum(d['weight'] for d in output_drums)
                has_drum_gc = any(d['lm_gc'] > 0 or d['ma_gc'] > 0 for d in output_drums)
                
                avg_lm_gc = 0.0
                avg_ma_gc = 0.0
                if total_out_weight > 0 and has_drum_gc:
                    avg_lm_gc = sum(d['weight'] * d['lm_gc'] for d in output_drums) / total_out_weight
                    avg_ma_gc = sum(d['weight'] * d['ma_gc'] for d in output_drums) / total_out_weight
                
                ma_conversion = 0.0
                heptane_loss = 0.0
                process_loss_pct = 0.0
                if final_input > 0:
                    process_loss_pct = ((final_input - final_output) / final_input) * 100.0
                
                if process_type == 'ENZYME_RXN':
                    process_loss_pct = 0.0
                    ma_conversion = calculate_ma_conversion(avg_lm_gc, avg_ma_gc)
                elif process_type in ['SRP', 'WASHING']:
                    heptane_loss = process_loss_pct
                
                # Write to Firestore
                batch_doc = {
                    "batch_id": batch_key,
                    "raw_batch_number": raw_batch_no,
                    "vessel_number": plant,
                    "process_type": process_type,
                    "start_date": start_dt.isoformat() if start_dt else None,
                    "end_date": end_dt.isoformat() if end_dt else None,
                    "total_input_weight": final_input,
                    "total_output_weight": final_output,
                    "process_loss_pct": process_loss_pct,
                    "enzyme_conversion_pct": ma_conversion,
                    "heptane_loss_pct": heptane_loss,
                    "final_gc": {
                        "lm_pct": avg_lm_gc,
                        "ma_pct": avg_ma_gc
                    },
                    "sop_compliant": True,
                    "source_file": filename,
                    "last_updated": datetime.utcnow().isoformat()
                }
                
                batches_ref.document(batch_key).set(clean_dict(batch_doc), merge=True)
                batches_count += 1
                
                # Upload Drums
                for d_idx, d in enumerate(input_drums):
                    drum_str = d['drum_no'] or f"drum_{d_idx}"
                    doc_id = f"{batch_key}_INPUT_{drum_str}".replace('/', '-').replace('#', '_')
                    drums_ref.document(doc_id).set(clean_dict({
                        "batch_id": batch_key,
                        "stage": "INPUT",
                        "material_desc": d['desc'],
                        "drum_number": d['drum_no'],
                        "drum_weight": d['weight']
                    }), merge=True)
                for d_idx, d in enumerate(output_drums):
                    drum_str = d['drum_no'] or f"drum_{d_idx}"
                    doc_id = f"{batch_key}_OUTPUT_{drum_str}".replace('/', '-').replace('#', '_')
                    drums_ref.document(doc_id).set(clean_dict({
                        "batch_id": batch_key,
                        "stage": "OUTPUT",
                        "material_desc": d['desc'],
                        "drum_number": d['drum_no'],
                        "drum_weight": d['weight'],
                        "lm_gc": d['lm_gc'],
                        "ma_gc": d['ma_gc'],
                        "hpt_gc": d['hpt_gc']
                    }), merge=True)
                
                i = j
                continue
            i += 1
            
    wb.close()
    return batches_count


def get_first_nonzero(row_data, indices, s):
    for idx in indices:
        if idx + s < len(row_data):
            val = safe_float(row_data[idx + s])
            if val > 0: return val
    return 0.0


def process_flavour_fragrance_file(file_content, filename):
    wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
    drums_ref = db.collection('drums')
    
    # Check if MN-MLQ layout
    collection_name = 'mn_mlq_drums' if 'MN-MLQ' in filename else 'drums'
    target_ref = db.collection(collection_name)
    
    drums_count = 0
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: continue
        
        header_row = rows[0]
        batch_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('BATCH' in val.upper() or 'BATCH NO' in val.upper())), None)
        drum_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DR NO' in val.upper() or 'DRUM' in val.upper())), None)
        weight_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('WT' in val.upper() or 'WEIGHT' in val.upper() or 'DRUM/ WT.' in val.upper())), None)
        desc_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DESC' in val.upper() or 'MATERIAL' in val.upper())), None)
        date_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'DATE' in val.upper()), None)
        
        lm_cols = [idx for idx, val in enumerate(header_row) if is_lm_header(val)]
        ma_cols = [idx for idx, val in enumerate(header_row) if is_ma_header(val)]
        hpt_cols = [idx for idx, val in enumerate(header_row) if is_hpt_header(val)]
        
        if batch_col is None or weight_col is None:
            continue
            
        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or all(v is None for v in row): continue
            
            batch_val = row[batch_col]
            if not batch_val or pd.isna(batch_val) or str(batch_val).strip() == '':
                continue
                
            plant, core_id = get_plant_and_core_id(batch_val)
            if not plant or not core_id:
                continue
            batch_key = f"{plant}_{core_id}"
            
            drum_no = str(row[drum_col]).strip() if drum_col is not None and row[drum_col] else "UNKNOWN"
            weight = safe_float(row[weight_col])
            desc = str(row[desc_col]).strip().upper() if desc_col is not None and row[desc_col] else "UNKNOWN"
            
            p_date = row[date_col] if date_col is not None else None
            p_date_str = p_date.isoformat() if isinstance(p_date, datetime) else str(p_date) if p_date else None
            
            lm_gc = safe_float(row[lm_cols[0]]) if lm_cols and row[lm_cols[0]] else 0.0
            ma_gc = safe_float(row[ma_cols[0]]) if ma_cols and row[ma_cols[0]] else 0.0
            hpt_gc = safe_float(row[hpt_cols[0]]) if hpt_cols and row[hpt_cols[0]] else 0.0
            
            doc_id = f"{batch_key}_OUTPUT_{drum_no}".replace('/', '-').replace('#', '_')
            
            target_ref.document(doc_id).set(clean_dict({
                "batch_id": batch_key,
                "stage": "OUTPUT",
                "material_desc": desc,
                "drum_number": drum_no,
                "drum_weight": weight,
                "lm_gc": lm_gc,
                "ma_gc": ma_gc,
                "hpt_gc": hpt_gc,
                "production_date": p_date_str
            }), merge=True)
            drums_count += 1
            
    wb.close()
    return drums_count


def process_production_logs(file_content, filename):
    dfs = pd.read_excel(file_content, sheet_name=None)
    drums_ref = db.collection('drums')
    drums_count = 0
    
    for sh, df in dfs.items():
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        batch_col = next((c for c in df.columns if 'batch' in c), None)
        drum_col = next((c for c in df.columns if 'dr no' in c or 'drum no' in c), None)
        weight_col = next((c for c in df.columns if 'wt' in c and 'drum' in c or 'drum/ wt.' in c), None)
        stage_col = next((c for c in df.columns if 'stap' in c or 'recycle' in c or 'stage' in c), None)
        desc_col = next((c for c in df.columns if 'description' in c or 'material' in c), None)
        date_col = next((c for c in df.columns if 'date' in c), None)
        
        if not batch_col or not weight_col:
            continue
            
        for _, row in df.iterrows():
            batch_val = row[batch_col]
            if pd.isna(batch_val) or str(batch_val).strip() == '' or str(batch_val).strip().lower() == 'nan':
                continue
                
            plant, core_id = get_plant_and_core_id(batch_val)
            if not plant or not core_id:
                continue
            batch_key = f"{plant}_{core_id}"
            
            drum_no = str(row[drum_col]).strip() if drum_col and pd.notna(row[drum_col]) else "UNKNOWN"
            weight = safe_float(row[weight_col])
            
            stage_str = str(row[stage_col]).upper() if stage_col and pd.notna(row[stage_col]) else ""
            is_output = 'RECYCLE' in stage_str or 'WIP' in stage_str or 'OUTPUT' in stage_str or 'FAA' in drum_no
            stage = "OUTPUT" if is_output else "INPUT"
            
            desc = str(row[desc_col]).strip().upper() if desc_col and pd.notna(row[desc_col]) else "UNKNOWN"
            p_date = row[date_col]
            p_date_str = p_date.isoformat() if isinstance(p_date, datetime) else str(p_date) if pd.notna(p_date) else None
                
            doc_id = f"{batch_key}_{stage}_{drum_no}".replace('/', '-').replace('#', '_')
            
            drums_ref.document(doc_id).set(clean_dict({
                "batch_id": batch_key,
                "stage": stage,
                "material_desc": desc,
                "drum_number": drum_no,
                "drum_weight": weight,
                "production_date": p_date_str
            }), merge=True)
            drums_count += 1
            
    return drums_count


def process_gc_file(file_content, filename):
    df = pd.read_excel(file_content, header=None)
    batches_ref = db.collection('batches')
    current_batch_key = None
    current_headers = None
    batches_count = 0
    
    for idx, row in df.iterrows():
        col0 = row.iloc[0]
        if pd.notna(col0) and isinstance(col0, str) and ('#' in col0 or 'P-' in col0 or 'V-' in col0):
            raw_batch_id = str(col0).strip().upper()
            plant, core_id = get_plant_and_core_id(raw_batch_id)
            if plant and core_id:
                current_batch_key = f"{plant}_{core_id}"
            else:
                current_batch_key = None
            continue
            
        if current_batch_key:
            if pd.notna(col0) and str(col0).strip().upper() == 'DATE':
                current_headers = [str(x).strip() if pd.notna(x) else f"Col_{i}" for i, x in enumerate(row.values)]
                continue
                
            if pd.notna(col0) and current_headers is not None:
                row_str = ' '.join([str(x).upper() for x in row.values if pd.notna(x)])
                is_feed = 'FEED' in row_str
                is_final = 'FINAL' in row_str and not is_feed
                
                if is_feed or is_final:
                    gc_data = {}
                    for i, h in enumerate(current_headers):
                        if not h.startswith("Col_") and h != 'DATE' and h != 'TIME':
                            val = row.values[i]
                            if pd.notna(val): gc_data[h] = safe_float(val)
                    
                    batch_doc_ref = batches_ref.document(current_batch_key)
                    doc_snap = batch_doc_ref.get()
                    
                    field_name = 'feed_gc' if is_feed else 'final_gc'
                    
                    lm_key = next((k for k in gc_data if is_lm_header(k)), None)
                    ma_key = next((k for k in gc_data if is_ma_header(k)), None)
                    hpt_key = next((k for k in gc_data if is_hpt_header(k)), None)
                    
                    standard_gc = {}
                    if lm_key: standard_gc['lm_pct'] = gc_data[lm_key]
                    if ma_key: standard_gc['ma_pct'] = gc_data[ma_key]
                    if hpt_key: standard_gc['heptane_pct'] = gc_data[hpt_key]
                    
                    if doc_snap.exists:
                        batch_doc_ref.update({
                            field_name: standard_gc,
                            f"{field_name}_full": clean_dict(gc_data)
                        })
                        if not is_feed:
                            snap_data = doc_snap.to_dict()
                            if snap_data.get('process_type') == 'ENZYME_RXN':
                                avg_lm = standard_gc.get('lm_pct', 0.0)
                                avg_ma = standard_gc.get('ma_pct', 0.0)
                                ma_conv = calculate_ma_conversion(avg_lm, avg_ma)
                                batch_doc_ref.update({"enzyme_conversion_pct": ma_conv})
                    else:
                        batch_doc_ref.set({
                            "batch_id": current_batch_key,
                            "raw_batch_number": raw_batch_id,
                            "vessel_number": plant,
                            field_name: standard_gc,
                            f"{field_name}_full": clean_dict(gc_data),
                            "last_updated": datetime.utcnow().isoformat()
                        }, merge=True)
                    batches_count += 1
                else:
                    gc_data = {}
                    for i, h in enumerate(current_headers):
                        if not h.startswith("Col_") and h != 'DATE' and h != 'TIME':
                            val = row.values[i]
                            if pd.notna(val): gc_data[h] = safe_float(val)
                    
                    sample_name = str(col0).strip()
                    doc_id = f"{sample_name}_{idx}".replace('/', '-').replace('#', '_')
                    batches_ref.document(current_batch_key).collection('timewise_gc').document(doc_id).set({
                        "sample": sample_name,
                        "date": str(row.values[0]) if pd.notna(row.values[0]) else None,
                        "time": str(row.values[1]) if pd.notna(row.values[1]) else None,
                        "gc_values": clean_dict(gc_data)
                    })
                    
    return batches_count
