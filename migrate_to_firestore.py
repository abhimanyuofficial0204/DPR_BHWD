import os
import re
import json
import math
import warnings
import openpyxl
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
from google.oauth2 import service_account

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 1. Initialize Firestore Database
print("Initializing Firestore Connection...")
CREDENTIAL_PATH = "firebase-adminsdk.json"

if not os.path.exists(CREDENTIAL_PATH):
    print(f"Error: {CREDENTIAL_PATH} not found!")
    exit(1)

credentials_obj = service_account.Credentials.from_service_account_file(CREDENTIAL_PATH)
db = firestore.Client(credentials=credentials_obj, database="prod123")
print("Connected to Firestore successfully.")

# Helper to clean dict of NaNs for Firestore
def clean_dict(d):
    clean = {}
    for k, v in d.items():
        if pd.isna(v) or v is None:
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        if isinstance(v, datetime):
            clean[k] = v.isoformat()
        else:
            clean[k] = v
    return clean

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace('%', ''))
    except:
        return default

# Extraction logic for Core ID and Plant
def get_plant_and_core_id(batch_str, default_plant=None):
    if not batch_str or pd.isna(batch_str) or str(batch_str).strip().upper() == 'NONE':
        return None, None
    s = str(batch_str).strip().upper()
    
    # Try extracting plant prefix (e.g. V-22, P-09, V22, P09, etc.)
    plant_match = re.search(r'([VP]-\d+)', s)
    plant = plant_match.group(1) if plant_match else default_plant
    
    # Extract core numeric sequence at the end of the string
    m = re.findall(r'\d+', s)
    core_id = None
    if m:
        core_id = str(int(m[-1])) # standardises '095' to '95'
        
    return plant, core_id

# Molecular weight conversion formula
def calculate_ma_conversion(output_lm_pct, output_ma_pct):
    if output_lm_pct == 0 and output_ma_pct == 0:
        return 0.0
    moles_ma = output_ma_pct / 198.30
    moles_lm = output_lm_pct / 156.27
    if (moles_ma + moles_lm) == 0:
        return 0.0
    return (moles_ma / (moles_ma + moles_lm)) * 100.0


SOURCE_DIR = "/Users/abhi/WorkBench/Sharp_Mint/DPR/New whatsapp download"

def migrate_dpr():
    filepath = os.path.join(SOURCE_DIR, "GLR_FY'26-27.xlsx")
    if not os.path.exists(filepath):
        print(f"DPR file not found: {filepath}")
        return
        
    print("\n--- Migrating DPR (Batches & Drums) ---")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    batches_ref = db.collection('batches')
    drums_ref = db.collection('drums')
    
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        
        # Build column indices map
        header_row = rows[0] if rows else []
        lm_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and ('L-MENTHOL' in col.upper() or col.upper() == 'LM')][:1]
        ma_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and 'MENTHYL ACETATE' in col.upper()][:1]
        hpt_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and 'HEPTANE' in col.upper()][:1]
        
        i = 0
        while i < len(rows):
            r = rows[i]
            if r and len(r) > 1 and (r[0] == 'Batch Size' or r[1] == 'Batch Size'):
                shift = 1 if r[1] == 'Batch Size' else 0
                r1 = rows[i]
                r2 = rows[i+1] if i+1 < len(rows) else [None]*150
                
                raw_batch_no = str(r1[9 + shift] or r2[9 + shift] or '').strip()
                plant, core_id = get_plant_and_core_id(raw_batch_no, default_plant=sh)
                
                if not core_id or not plant:
                    i += 1
                    continue
                    
                batch_key = f"{plant}_{core_id}"
                
                header_input_weight = safe_float(r1[1 + shift])
                
                # Locate Output Qty
                header_output_weight = 0.0
                if sh != 'V-24':
                    for r_hdr in [r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                        if r_hdr:
                            for c_idx, cell_val in enumerate(r_hdr):
                                if str(cell_val).strip().lower() == 'output qty' and c_idx + 1 < len(r_hdr):
                                    header_output_weight = safe_float(r_hdr[c_idx + 1])
                                    break
                                    
                # Locate Start & End Dates
                start_dt = None
                end_dt = None
                for r_hdr in [rows[i-1] if i>0 else [], r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                    if r_hdr:
                        for c_idx, cell_val in enumerate(r_hdr):
                            val_str = str(cell_val).strip().lower()
                            if 'start date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime): start_dt = candidate
                            elif 'end date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime): end_dt = candidate
                
                # Parse drum rows
                input_drums = []
                output_drums = []
                j = i + 1
                while j < len(rows) and j < i + 4:
                    dr_row = rows[j]
                    if not dr_row: break
                    c0 = str(dr_row[0] or '').strip().lower()
                    c1 = str(dr_row[1] or '').strip().lower()
                    if 'batch size' in c0 or 'batch size' in c1:
                        j += 1
                        continue
                    break
                
                has_trf_in_process = False
                while j < len(rows):
                    dr_row = rows[j]
                    if not dr_row or (len(dr_row) > 1 and (dr_row[0] == 'Batch Size' or dr_row[1] == 'Batch Size')):
                        break
                    if all(val is None or str(val).strip() == '' for val in dr_row):
                        break
                        
                    # Input drum
                    in_desc = str(dr_row[0 + shift] or '').strip().upper()
                    in_dr_no = str(dr_row[1 + shift] or '').strip()
                    in_weight = safe_float(dr_row[3 + shift] if 3+shift < len(dr_row) else 0)
                    if in_weight > 0 or in_desc != '':
                        input_drums.append({'desc': in_desc, 'drum_no': in_dr_no, 'weight': in_weight})
                        
                    # Output drum
                    out_weight_val = dr_row[18 + shift] if 18+shift < len(dr_row) else 0
                    if str(out_weight_val).strip().upper() == 'TRF IN PROCESS':
                        has_trf_in_process = True
                        
                    out_desc_col = 26 + shift
                    out_desc = str(dr_row[out_desc_col] if out_desc_col < len(dr_row) else '').strip().upper()
                    if out_desc == '' and 11+shift < len(dr_row):
                        out_desc = str(dr_row[11+shift] or '').strip().upper()
                        
                    out_weight = safe_float(out_weight_val)
                    out_dr_no = ''
                    if 15+shift < len(dr_row) and dr_row[15+shift]:
                        out_dr_no = str(dr_row[15+shift]).strip()
                    elif 11+shift < len(dr_row) and dr_row[11+shift]:
                        out_dr_no = str(dr_row[11+shift]).strip()
                        
                    def get_first_nonzero(row_data, indices, s):
                        for idx in indices:
                            if idx + s < len(row_data):
                                val = safe_float(row_data[idx + s])
                                if val > 0: return val
                        return 0.0
                        
                    out_lm = get_first_nonzero(dr_row, lm_cols, shift)
                    out_ma = get_first_nonzero(dr_row, ma_cols, shift)
                    out_hpt = get_first_nonzero(dr_row, hpt_cols, shift)
                    
                    if out_weight > 0 or out_desc != '':
                        output_drums.append({
                            'desc': out_desc,
                            'drum_no': out_dr_no,
                            'weight': out_weight,
                            'lm_gc': out_lm,
                            'ma_gc': out_ma,
                            'hpt_gc': out_hpt
                        })
                    j += 1
                
                # Determine Process
                feed_types = set([d['desc'] for d in input_drums if d['desc']])
                output_types = set([d['desc'] for d in output_drums if d['desc'] and d['desc'] != 'WATER'])
                raw_process = str(r1[6 + shift] or '').strip().upper()
                
                process_type = None
                if 'DLM' in feed_types: process_type = 'ENZYME_RXN'
                elif 'ADH' in feed_types: process_type = 'SRP'
                elif 'HPT' in feed_types and 'NHPT' in output_types: process_type = 'WASHING'
                elif 'DMM' in feed_types: process_type = 'DISTILLATION'
                
                if not process_type:
                    if 'GLR' in raw_process or 'ENZYME' in raw_process: process_type = 'ENZYME_RXN'
                    elif 'SRP' in raw_process: process_type = 'SRP'
                    elif 'WASH' in raw_process: process_type = 'WASHING'
                    else: process_type = 'UNKNOWN'
                
                sum_in_drums = sum(d['weight'] for d in input_drums)
                sum_out_drums = sum(d['weight'] for d in output_drums)
                
                final_input = sum_in_drums if sum_in_drums > 0 else header_input_weight
                final_output = sum_out_drums if sum_out_drums > 0 else header_output_weight
                if process_type == 'ENZYME_RXN' and final_output == 0:
                    final_output = final_input
                
                # GC Averages
                total_out_weight = sum(d['weight'] for d in output_drums)
                has_drum_gc = any(d['lm_gc'] > 0 or d['ma_gc'] > 0 for d in output_drums)
                
                avg_lm_gc = 0.0
                avg_ma_gc = 0.0
                if total_out_weight > 0 and has_drum_gc:
                    avg_lm_gc = sum(d['weight'] * d['lm_gc'] for d in output_drums) / total_out_weight
                    avg_ma_gc = sum(d['weight'] * d['ma_gc'] for d in output_drums) / total_out_weight
                
                ma_conversion = 0.0
                heptane_loss = 0.0
                process_loss_pct = 0.0
                if final_input > 0:
                    process_loss_pct = ((final_input - final_output) / final_input) * 100.0
                
                if process_type == 'ENZYME_RXN':
                    process_loss_pct = 0.0
                    ma_conversion = calculate_ma_conversion(avg_lm_gc, avg_ma_gc)
                elif process_type in ['SRP', 'WASHING']:
                    heptane_loss = process_loss_pct
                
                # Set Firestore document
                batch_doc = {
                    "batch_id": batch_key,
                    "raw_batch_number": raw_batch_no,
                    "vessel_number": plant,
                    "process_type": process_type,
                    "start_date": start_dt.isoformat() if start_dt else None,
                    "end_date": end_dt.isoformat() if end_dt else None,
                    "total_input_weight": final_input,
                    "total_output_weight": final_output,
                    "process_loss_pct": process_loss_pct,
                    "enzyme_conversion_pct": ma_conversion,
                    "heptane_loss_pct": heptane_loss,
                    "final_gc": {
                        "lm_pct": avg_lm_gc,
                        "ma_pct": avg_ma_gc
                    },
                    "sop_compliant": True,
                    "source_file": "GLR_FY'26-27.xlsx",
                    "last_updated": datetime.utcnow().isoformat()
                }
                
                batches_ref.document(batch_key).set(clean_dict(batch_doc))
                
                # Upload Drums
                for d_idx, d in enumerate(input_drums):
                    drum_str = d['drum_no'] or f"drum_{d_idx}"
                    doc_id = f"{batch_key}_INPUT_{drum_str}".replace('/', '-').replace('#', '_')
                    drums_ref.document(doc_id).set(clean_dict({
                        "batch_id": batch_key,
                        "stage": "INPUT",
                        "material_desc": d['desc'],
                        "drum_number": d['drum_no'],
                        "drum_weight": d['weight']
                    }))
                for d_idx, d in enumerate(output_drums):
                    drum_str = d['drum_no'] or f"drum_{d_idx}"
                    doc_id = f"{batch_key}_OUTPUT_{drum_str}".replace('/', '-').replace('#', '_')
                    drums_ref.document(doc_id).set(clean_dict({
                        "batch_id": batch_key,
                        "stage": "OUTPUT",
                        "material_desc": d['desc'],
                        "drum_number": d['drum_no'],
                        "drum_weight": d['weight'],
                        "lm_gc": d['lm_gc'],
                        "ma_gc": d['ma_gc'],
                        "hpt_gc": d['hpt_gc']
                    }))
                
                i = j
                continue
            i += 1
            
    wb.close()
    print("DPR migration finished.")

def migrate_gc_file(filename):
    filepath = os.path.join(SOURCE_DIR, filename)
    if not os.path.exists(filepath):
        print(f"GC file not found: {filepath}")
        return
        
    print(f"\n--- Migrating GC {filename} ---")
    df = pd.read_excel(filepath, header=None)
    
    batches_ref = db.collection('batches')
    current_batch_key = None
    current_headers = None
    
    for idx, row in df.iterrows():
        col0 = row.iloc[0]
        if pd.notna(col0) and isinstance(col0, str) and ('#' in col0 or 'P-' in col0 or 'V-' in col0):
            raw_batch_id = str(col0).strip().upper()
            plant, core_id = get_plant_and_core_id(raw_batch_id)
            if plant and core_id:
                current_batch_key = f"{plant}_{core_id}"
            else:
                current_batch_key = None
            continue
            
        if current_batch_key:
            if pd.notna(col0) and str(col0).strip().upper() == 'DATE':
                current_headers = [str(x).strip() if pd.notna(x) else f"Col_{i}" for i, x in enumerate(row.values)]
                continue
                
            if pd.notna(col0) and current_headers is not None:
                row_str = ' '.join([str(x).upper() for x in row.values if pd.notna(x)])
                is_feed = 'FEED' in row_str
                is_final = 'FINAL' in row_str and not is_feed
                
                if is_feed or is_final:
                    # Construct GC object
                    gc_data = {}
                    for i, h in enumerate(current_headers):
                        if not h.startswith("Col_") and h != 'DATE' and h != 'TIME':
                            val = row.values[i]
                            if pd.notna(val):
                                gc_data[h] = safe_float(val)
                    
                    # Update firestore batch
                    batch_doc_ref = batches_ref.document(current_batch_key)
                    # Check if batch exists, if not, create it
                    doc_snap = batch_doc_ref.get()
                    
                    field_name = 'feed_gc' if is_feed else 'final_gc'
                    
                    # Ensure LM and MA are mapped to standardized keys
                    lm_key = next((k for k in gc_data if 'LM' in k.upper() or 'L-MENTHOL' in k.upper()), None)
                    ma_key = next((k for k in gc_data if 'MENTHYL ACETATE' in k.upper() or 'MA' in k.upper()), None)
                    hpt_key = next((k for k in gc_data if 'HEPTANE' in k.upper() or 'HPT' in k.upper()), None)
                    
                    standard_gc = {}
                    if lm_key: standard_gc['lm_pct'] = gc_data[lm_key]
                    if ma_key: standard_gc['ma_pct'] = gc_data[ma_key]
                    if hpt_key: standard_gc['heptane_pct'] = gc_data[hpt_key]
                    
                    if doc_snap.exists:
                        batch_doc_ref.update({
                            field_name: standard_gc,
                            f"{field_name}_full": clean_dict(gc_data)
                        })
                        # Re-calculate conversion % if final_gc LM/MA is updated
                        if not is_feed:
                            snap_data = doc_snap.to_dict()
                            if snap_data.get('process_type') == 'ENZYME_RXN':
                                avg_lm = standard_gc.get('lm_pct', 0.0)
                                avg_ma = standard_gc.get('ma_pct', 0.0)
                                ma_conv = calculate_ma_conversion(avg_lm, avg_ma)
                                batch_doc_ref.update({"enzyme_conversion_pct": ma_conv})
                    else:
                        # Write base batch document
                        batch_doc_ref.set({
                            "batch_id": current_batch_key,
                            "raw_batch_number": raw_batch_id,
                            "vessel_number": plant,
                            field_name: standard_gc,
                            f"{field_name}_full": clean_dict(gc_data),
                            "last_updated": datetime.utcnow().isoformat()
                        }, merge=True)
                else:
                    # Intermediate/timewise sample
                    # We store it in a sub-collection 'timewise_gc'
                    gc_data = {}
                    for i, h in enumerate(current_headers):
                        if not h.startswith("Col_") and h != 'DATE' and h != 'TIME':
                            val = row.values[i]
                            if pd.notna(val): gc_data[h] = safe_float(val)
                    
                    sample_name = str(col0).strip()
                    doc_id = f"{sample_name}_{idx}"
                    batches_ref.document(current_batch_key).collection('timewise_gc').document(doc_id).set({
                        "sample": sample_name,
                        "date": str(row.values[0]) if pd.notna(row.values[0]) else None,
                        "time": str(row.values[1]) if pd.notna(row.values[1]) else None,
                        "gc_values": clean_dict(gc_data)
                    })

def migrate_production_logs():
    filepath = os.path.join(SOURCE_DIR, "New Pro 26-27.xlsx")
    if not os.path.exists(filepath):
        print(f"Production logs not found: {filepath}")
        return
        
    print("\n--- Migrating Production Logs (Drums) ---")
    # Read sheet 'SYN' or 'NAT.'
    dfs = pd.read_excel(filepath, sheet_name=None)
    drums_ref = db.collection('drums')
    
    for sh, df in dfs.items():
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        batch_col = next((c for c in df.columns if 'batch' in c), None)
        drum_col = next((c for c in df.columns if 'dr no' in c or 'drum no' in c), None)
        weight_col = next((c for c in df.columns if 'wt' in c and 'drum' in c or 'drum/ wt.' in c), None)
        stage_col = next((c for c in df.columns if 'stap' in c or 'recycle' in c or 'stage' in c), None)
        desc_col = next((c for c in df.columns if 'description' in c or 'material' in c), None)
        date_col = next((c for c in df.columns if 'date' in c), None)
        
        if not batch_col or not weight_col:
            continue
            
        for _, row in df.iterrows():
            batch_val = row[batch_col]
            if pd.isna(batch_val) or str(batch_val).strip() == '' or str(batch_val).strip().lower() == 'nan':
                continue
                
            plant, core_id = get_plant_and_core_id(batch_val)
            if not plant or not core_id:
                continue
            batch_key = f"{plant}_{core_id}"
            
            drum_no = str(row[drum_col]).strip() if drum_col and pd.notna(row[drum_col]) else "UNKNOWN"
            weight = safe_float(row[weight_col])
            
            stage_str = str(row[stage_col]).upper() if stage_col and pd.notna(row[stage_col]) else ""
            is_output = 'RECYCLE' in stage_str or 'WIP' in stage_str or 'OUTPUT' in stage_str or 'FAA' in drum_no
            stage = "OUTPUT" if is_output else "INPUT"
            
            desc = str(row[desc_col]).strip().upper() if desc_col and pd.notna(row[desc_col]) else "UNKNOWN"
            p_date = row[date_col]
            if isinstance(p_date, datetime):
                p_date_str = p_date.isoformat()
            else:
                p_date_str = str(p_date) if pd.notna(p_date) else None
                
            doc_id = f"{batch_key}_{stage}_{drum_no}".replace('/', '-').replace('#', '_')
            
            # Upsert into Firestore
            drums_ref.document(doc_id).set(clean_dict({
                "batch_id": batch_key,
                "stage": stage,
                "material_desc": desc,
                "drum_number": drum_no,
                "drum_weight": weight,
                "production_date": p_date_str
            }), merge=True)

def migrate_flavour_fragrance_file(filename, is_flavour=True):
    filepath = os.path.join(SOURCE_DIR, filename)
    if not os.path.exists(filepath):
        print(f"Flavour/Fragrance file not found: {filepath}")
        return
        
    print(f"\n--- Migrating {filename} ---")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    drums_ref = db.collection('drums')
    
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: continue
        
        # Parse headers
        header_row = rows[0]
        col_map = {}
        for idx, val in enumerate(header_row):
            if isinstance(val, str):
                col_map[val.strip().upper()] = idx
                
        batch_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('BATCH' in val.upper() or 'BATCH NO' in val.upper())), None)
        drum_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DR NO' in val.upper() or 'DRUM' in val.upper())), None)
        weight_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('WT' in val.upper() or 'WEIGHT' in val.upper() or 'DRUM/ WT.' in val.upper())), None)
        desc_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DESC' in val.upper() or 'MATERIAL' in val.upper())), None)
        date_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'DATE' in val.upper()), None)
        
        # Chemical compounds GC
        lm_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('LM' in val.upper() or 'L-MENTHOL' in val.upper())]
        ma_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'MENTHYL ACETATE' in val.upper()]
        hpt_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'HEPTANE' in val.upper()]
        
        if batch_col is None or weight_col is None:
            continue
            
        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or all(v is None for v in row): continue
            
            batch_val = row[batch_col]
            if not batch_val or pd.isna(batch_val) or str(batch_val).strip() == '':
                continue
                
            plant, core_id = get_plant_and_core_id(batch_val)
            if not plant or not core_id:
                continue
            batch_key = f"{plant}_{core_id}"
            
            drum_no = str(row[drum_col]).strip() if drum_col is not None and row[drum_col] else "UNKNOWN"
            weight = safe_float(row[weight_col])
            desc = str(row[desc_col]).strip().upper() if desc_col is not None and row[desc_col] else "UNKNOWN"
            
            p_date = row[date_col] if date_col is not None else None
            p_date_str = p_date.isoformat() if isinstance(p_date, datetime) else str(p_date) if p_date else None
            
            # GC percentages
            lm_gc = safe_float(row[lm_cols[0]]) if lm_cols and row[lm_cols[0]] else 0.0
            ma_gc = safe_float(row[ma_cols[0]]) if ma_cols and row[ma_cols[0]] else 0.0
            hpt_gc = safe_float(row[hpt_cols[0]]) if hpt_cols and row[hpt_cols[0]] else 0.0
            
            doc_id = f"{batch_key}_OUTPUT_{drum_no}".replace('/', '-').replace('#', '_')
            
            drums_ref.document(doc_id).set(clean_dict({
                "batch_id": batch_key,
                "stage": "OUTPUT",
                "material_desc": desc,
                "drum_number": drum_no,
                "drum_weight": weight,
                "lm_gc": lm_gc,
                "ma_gc": ma_gc,
                "hpt_gc": hpt_gc,
                "production_date": p_date_str
            }), merge=True)
            
    wb.close()
    print(f"{filename} migration finished.")

def migrate_mn_mlq():
    filename = "MN-MLQ-25-26.xlsx"
    filepath = os.path.join(SOURCE_DIR, filename)
    if not os.path.exists(filepath):
        print(f"MN-MLQ file not found: {filepath}")
        return
        
    print(f"\n--- Migrating {filename} ---")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    mn_mlq_ref = db.collection('mn_mlq_drums')
    
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: continue
        
        header_row = rows[0]
        batch_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('BATCH' in val.upper() or 'BATCH NO' in val.upper())), None)
        drum_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DR NO' in val.upper() or 'DRUM' in val.upper())), None)
        weight_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('WT' in val.upper() or 'WEIGHT' in val.upper() or 'DRUM/ WT.' in val.upper())), None)
        desc_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('DESC' in val.upper() or 'MATERIAL' in val.upper())), None)
        date_col = next((idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'DATE' in val.upper()), None)
        
        lm_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and ('LM' in val.upper() or 'L-MENTHOL' in val.upper())]
        ma_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'MENTHYL ACETATE' in val.upper()]
        hpt_cols = [idx for idx, val in enumerate(header_row) if isinstance(val, str) and 'HEPTANE' in val.upper()]
        
        if batch_col is None or weight_col is None:
            continue
            
        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or all(v is None for v in row): continue
            
            batch_val = row[batch_col]
            if not batch_val or pd.isna(batch_val) or str(batch_val).strip() == '':
                continue
                
            plant, core_id = get_plant_and_core_id(batch_val)
            if not plant or not core_id:
                continue
            batch_key = f"{plant}_{core_id}"
            
            drum_no = str(row[drum_col]).strip() if drum_col is not None and row[drum_col] else "UNKNOWN"
            weight = safe_float(row[weight_col])
            desc = str(row[desc_col]).strip().upper() if desc_col is not None and row[desc_col] else "UNKNOWN"
            
            p_date = row[date_col] if date_col is not None else None
            p_date_str = p_date.isoformat() if isinstance(p_date, datetime) else str(p_date) if p_date else None
            
            lm_gc = safe_float(row[lm_cols[0]]) if lm_cols and row[lm_cols[0]] else 0.0
            ma_gc = safe_float(row[ma_cols[0]]) if ma_cols and row[ma_cols[0]] else 0.0
            hpt_gc = safe_float(row[hpt_cols[0]]) if hpt_cols and row[hpt_cols[0]] else 0.0
            
            doc_id = f"{batch_key}_OUTPUT_{drum_no}".replace('/', '-').replace('#', '_')
            
            mn_mlq_ref.document(doc_id).set(clean_dict({
                "batch_id": batch_key,
                "stage": "OUTPUT",
                "material_desc": desc,
                "drum_number": drum_no,
                "drum_weight": weight,
                "lm_gc": lm_gc,
                "ma_gc": ma_gc,
                "hpt_gc": hpt_gc,
                "production_date": p_date_str
            }), merge=True)
            
    wb.close()
    print("MN-MLQ migration finished.")

if __name__ == "__main__":
    # Execute full migration chain
    print("Starting full Firestore Migration...")
    migrate_dpr()
    migrate_gc_file("GLR-26-27.xlsx")
    migrate_gc_file("RXN-26-27.xlsx")
    migrate_production_logs()
    migrate_flavour_fragrance_file("FLAOUR  PROD-- -26-27.xlsx", is_flavour=True)
    migrate_flavour_fragrance_file("FRAGRANCE PROD.-2026-27.xlsx", is_flavour=False)
    migrate_mn_mlq()
    print("Full Migration successfully finished!")
