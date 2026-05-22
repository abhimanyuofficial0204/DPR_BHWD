import openpyxl
import pandas as pd
from datetime import datetime
import os
import sqlite3
from sop_config import PHASE_2_SOP
from db import get_connection

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace('%', ''))
    except:
        return default

def calculate_ma_conversion(output_lm_pct, output_ma_pct):
    if output_lm_pct == 0 and output_ma_pct == 0:
        return 0.0
    moles_ma = output_ma_pct / 198.30
    moles_lm = output_lm_pct / 156.27
    if (moles_ma + moles_lm) == 0:
        return 0.0
    return (moles_ma / (moles_ma + moles_lm)) * 100.0

def extract_and_upsert_dpr(filepath):
    print(f"Extracting data from {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    conn = get_connection()
    cursor = conn.cursor()
    
    for sh in wb.sheetnames:
        sheet = wb[sh]
        rows = list(sheet.iter_rows(values_only=True))
        
        # Build column index map from header (first few rows)
        col_map = {}
        header_row = rows[0] if rows else []
        for idx, col_name in enumerate(header_row):
            if isinstance(col_name, str):
                col_map[col_name.strip().upper()] = idx
        
        # Find all occurrences of GC columns to handle FEED vs FINAL GC sections
        # We limit to the FIRST occurrence [:1] to prevent accidentally picking up identical column names 
        # further to the right which are often used for calculated weights or summaries.
        lm_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and ('L-MENTHOL' in col.upper() or col.upper() == 'LM')][:1]
        ma_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and 'MENTHYL ACETATE' in col.upper()][:1]
        hpt_cols = [idx for idx, col in enumerate(header_row) if isinstance(col, str) and 'HEPTANE' in col.upper()][:1]
        
        i = 0
        while i < len(rows):
            r = rows[i]
            if r and len(r) > 1 and (r[0] == 'Batch Size' or r[1] == 'Batch Size'):
                shift = 1 if r[1] == 'Batch Size' else 0
                r1 = rows[i]
                r2 = rows[i+1] if i+1 < len(rows) else [None]*150
                
                batch_no = str(r1[9 + shift] or r2[9 + shift] or '').strip()
                
                header_input_weight = safe_float(r1[1 + shift])
                
                # Dynamically locate 'OutPut Qty' in the header block
                header_output_weight = 0.0
                if sh != 'V-24':
                    for r_hdr in [r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                        if r_hdr:
                            for c_idx, cell_val in enumerate(r_hdr):
                                if str(cell_val).strip().lower() == 'output qty' and c_idx + 1 < len(r_hdr):
                                    header_output_weight = safe_float(r_hdr[c_idx + 1])
                                    break
                
                # Dynamically locate Start Date and End Date in the header block
                start_dt = None
                end_dt = None
                
                for r_hdr in [r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                    if r_hdr:
                        for c_idx, cell_val in enumerate(r_hdr):
                            val_str = str(cell_val).strip().lower()
                            if 'start date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime):
                                    start_dt = candidate
                            elif 'end date' in val_str and c_idx + 1 < len(r_hdr):
                                candidate = r_hdr[c_idx + 1]
                                if isinstance(candidate, datetime):
                                    end_dt = candidate
                
                def get_first_nonzero(row_data, indices, s):
                    for idx in indices:
                        if idx + s < len(row_data):
                            val = safe_float(row_data[idx + s])
                            if val > 0: return val
                    return 0.0

                # Header GCs (Fallback for Tank-to-Tank transfers)
                header_lm_gc = 0.0
                header_ma_gc = 0.0
                feed_lm_gc = 0.0
                feed_ma_gc = 0.0
                
                # Scan for 'FINAL' and 'FEED' in header rows to get the correct GC rows
                for hdr_row in [r1, r2, rows[i+2] if i+2 < len(rows) else []]:
                    if hdr_row:
                        # Find if any cell has 'FINAL' or 'FEED'
                        has_final = False
                        has_feed = False
                        for cell_val in hdr_row:
                            if isinstance(cell_val, str):
                                if 'FINAL' in cell_val.upper():
                                    has_final = True
                                elif 'FEED' in cell_val.upper():
                                    has_feed = True
                        if has_final:
                            header_lm_gc = get_first_nonzero(hdr_row, lm_cols, 0)
                            header_ma_gc = get_first_nonzero(hdr_row, ma_cols, 0)
                        if has_feed:
                            feed_lm_gc = get_first_nonzero(hdr_row, lm_cols, 0)
                            feed_ma_gc = get_first_nonzero(hdr_row, ma_cols, 0)
                
                raw_process = str(r1[6 + shift] or '').strip().upper()
                
                # Parse drum rows
                input_drums = []
                output_drums = []
                
                # Robustly skip the 2-3 row header block
                j = i + 1
                while j < len(rows) and j < i + 4:
                    dr_row = rows[j]
                    if not dr_row:
                        break
                    c0 = str(dr_row[0] or '').strip().lower()
                    c1 = str(dr_row[1] or '').strip().lower()
                    c2 = str(dr_row[2] or '').strip().lower()
                    
                    if 'batch size' in c0 or 'batch size' in c1:
                        j += 1
                        continue
                    if 'remaining material' in c2 or 'output qty' in c2 or 'loss' in c2:
                        j += 1
                        continue
                    break
                    
                # Determine if TRF IN PROCESS is in any output drum weight
                has_trf_in_process = False
                
                while j < len(rows):
                    dr_row = rows[j]
                    if not dr_row:
                        break
                    if len(dr_row) > 1 and (dr_row[0] == 'Batch Size' or dr_row[1] == 'Batch Size'):
                        break
                        
                    is_blank = all(val is None or str(val).strip() == '' for val in dr_row)
                    if is_blank:
                        break
                        
                    # Input drum: Col 0 is desc, Col 3 is weight
                    in_desc = str(dr_row[0 + shift] or '').strip().upper()
                    in_dr_no = str(dr_row[1 + shift] or '').strip()
                    in_weight = safe_float(dr_row[3 + shift] if 3+shift < len(dr_row) else 0)
                    if in_weight > 0 or in_desc != '':
                        input_drums.append({'desc': in_desc, 'drum_no': in_dr_no, 'weight': in_weight})
                        
                    # Output drum: Col 11 (or 26) is desc, Col 18 is weight
                    out_weight_val = dr_row[18 + shift] if 18+shift < len(dr_row) else 0
                    
                    if str(out_weight_val).strip().upper() == 'TRF IN PROCESS':
                        has_trf_in_process = True
                        
                    out_desc_col = 26 + shift
                    out_weight_col = 18 + shift
                    out_desc = str(dr_row[out_desc_col] if out_desc_col < len(dr_row) else '').strip().upper()
                    
                    # Also fallback to index 11 for description if index 26 is empty
                    if out_desc == '' and 11+shift < len(dr_row):
                        out_desc = str(dr_row[11+shift] or '').strip().upper()
                        
                    out_weight = safe_float(out_weight_val)
                    
                    out_dr_no = ''
                    if 15+shift < len(dr_row) and dr_row[15+shift]:
                        out_dr_no = str(dr_row[15+shift]).strip()
                    elif 11+shift < len(dr_row) and dr_row[11+shift]:
                        out_dr_no = str(dr_row[11+shift]).strip()
                    
                    # GC per drum (check all possible columns for the first non-zero value)
                    def get_first_nonzero(row_data, indices, s):
                        for idx in indices:
                            if idx + s < len(row_data):
                                val = safe_float(row_data[idx + s])
                                if val > 0: return val
                        return 0.0
                    
                    out_lm = get_first_nonzero(dr_row, lm_cols, shift)
                    out_ma = get_first_nonzero(dr_row, ma_cols, shift)
                    out_hpt = get_first_nonzero(dr_row, hpt_cols, shift)
                    
                    if out_weight > 0 or out_desc != '':
                        output_drums.append({
                            'desc': out_desc, 
                            'drum_no': out_dr_no,
                            'weight': out_weight,
                            'lm_gc': out_lm,
                            'ma_gc': out_ma,
                            'hpt_gc': out_hpt
                        })
                    
                    j += 1
                
                # Determine Process via Rule 2 & Header Fallback
                feed_types = set([d['desc'] for d in input_drums if d['desc']])
                output_types = set([d['desc'] for d in output_drums if d['desc'] and d['desc'] != 'WATER'])
                
                process_type = None
                if 'DLM' in feed_types:
                    process_type = 'ENZYME_RXN'
                elif 'ADH' in feed_types:
                    process_type = 'SRP'
                elif 'HPT' in feed_types and 'NHPT' in output_types:
                    process_type = 'WASHING'
                elif 'DMM' in feed_types:
                    process_type = 'DISTILLATION'
                
                # Fallback to header process if drum feed identification missed it
                if not process_type:
                    if 'GLR' in raw_process or 'ENZYME' in raw_process: process_type = 'ENZYME_RXN'
                    elif 'SRP' in raw_process: process_type = 'SRP'
                    elif 'DISTILL' in raw_process: process_type = 'DISTILLATION'
                    elif 'WASH' in raw_process: process_type = 'WASHING'
                    else:
                        i = j
                        continue # Skip if completely unrecognized
                        
                # Validation & Cross Check
                sum_in_drums = sum(d['weight'] for d in input_drums)
                sum_out_drums = sum(d['weight'] for d in output_drums)
                
                notes = []
                
                final_input = header_input_weight
                if abs(header_input_weight - sum_in_drums) > 5.0 and sum_in_drums > 0:
                    notes.append(f"Input Weight Mismatch: Header ({header_input_weight}) vs Drums ({sum_in_drums})")
                    final_input = sum_in_drums
                elif final_input == 0 and sum_in_drums > 0:
                    final_input = sum_in_drums
                    
                if sh == 'V-24':
                    final_output = sum_out_drums
                else:
                    final_output = header_output_weight
                    # Suppress output weight mismatch for ENZYME_RXN as batch transfers don't list weights
                    if not (process_type == 'ENZYME_RXN' and sum_out_drums < header_output_weight):
                        if abs(header_output_weight - sum_out_drums) > 5.0 and sum_out_drums > 0:
                            notes.append(f"Output Weight Mismatch: Header ({header_output_weight}) vs Drums ({sum_out_drums})")
                            final_output = sum_out_drums
                        elif final_output == 0 and sum_out_drums > 0:
                            final_output = sum_out_drums
                            
                if process_type == 'ENZYME_RXN' and final_output == 0:
                    final_output = final_input
                
                # Weighted Average GC calculations
                total_out_weight = sum(d['weight'] for d in output_drums)
                
                is_transfer = any('TRF' in d['desc'] or 'ST-' in d['desc'] for d in output_drums)
                has_drum_gc = any(d['lm_gc'] > 0 or d['ma_gc'] > 0 for d in output_drums)
                
                if total_out_weight > 0 and not is_transfer and has_drum_gc:
                    avg_lm_gc = sum(d['weight'] * d['lm_gc'] for d in output_drums) / total_out_weight
                    avg_ma_gc = sum(d['weight'] * d['ma_gc'] for d in output_drums) / total_out_weight
                else:
                    avg_lm_gc = header_lm_gc
                    avg_ma_gc = header_ma_gc
                    if is_transfer:
                        notes.append("Tank-to-Tank Transfer: Using Header GC")
                
                ma_conversion = 0.0
                heptane_loss = 0.0
                process_loss_pct = 0.0
                
                if final_input > 0:
                    process_loss_pct = ((final_input - final_output) / final_input) * 100.0
                    
                if process_type == 'ENZYME_RXN':
                    process_loss_pct = 0.0  # Enzyme reactions have no process loss
                
                sop_compliant = True
                
                if process_type == 'ENZYME_RXN':
                    if avg_lm_gc == 0 and avg_ma_gc == 0:
                        notes.append("GC not mentioned")
                    elif avg_ma_gc == 0:
                        notes.append("GC not mentioned")
                    else:
                        ma_conversion = calculate_ma_conversion(avg_lm_gc, avg_ma_gc)
                        min_c, max_c = PHASE_2_SOP['ENZYME_RXN']['target_conversion_LM_to_MA']
                        if not (min_c <= ma_conversion <= max_c):
                            sop_compliant = False
                            notes.append(f"MA Conversion {ma_conversion:.2f}% out of bounds ({min_c}-{max_c}%)")
                        
                elif process_type in ['SRP', 'WASHING']:
                    if has_trf_in_process:
                        heptane_loss = 0.0
                        notes.append("Heptane loss not calculated due to transfer")
                    else:
                        # Heptane loss is simply the overall process loss for these phases
                        heptane_loss = process_loss_pct
                        if process_type == 'SRP':
                            min_hl, max_hl = PHASE_2_SOP['SRP']['target_heptane_loss']
                            if not (min_hl <= heptane_loss <= max_hl):
                                sop_compliant = False
                                notes.append(f"Heptane Loss {heptane_loss:.2f}% out of bounds (max {max_hl}%)")
                        
                elif process_type == 'DISTILLATION':
                    min_pl, max_pl = PHASE_2_SOP['DISTILLATION']['target_process_loss']
                    if not (min_pl <= process_loss_pct <= max_pl):
                        sop_compliant = False
                        notes.append(f"Process Loss {process_loss_pct:.2f}% out of bounds ({min_pl}-{max_pl}%)")
                
                dev_notes = " | ".join(notes)
                initial_lm_pct = feed_lm_gc
                initial_ma_pct = feed_ma_gc
                final_lm_pct = avg_lm_gc
                final_ma_pct = avg_ma_gc

                # DB Insert
                cursor.execute('''
                    INSERT INTO DPR_Master (
                        Batch_Number, Process_Type, Start_Date, End_Date, Total_Input_Weight, Total_Output_Weight,
                        Process_Loss_Pct, LM_to_MA_Conversion_Pct, Heptane_Loss_Pct, Initial_LM_Pct, Initial_MA_Pct, Final_LM_Pct, Final_MA_Pct, SOP_Compliant, Deviation_Notes, Source_File
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(Batch_Number) DO UPDATE SET
                        Process_Type=excluded.Process_Type,
                        Start_Date=excluded.Start_Date,
                        End_Date=excluded.End_Date,
                        Total_Input_Weight=excluded.Total_Input_Weight,
                        Total_Output_Weight=excluded.Total_Output_Weight,
                        Process_Loss_Pct=excluded.Process_Loss_Pct,
                        LM_to_MA_Conversion_Pct=excluded.LM_to_MA_Conversion_Pct,
                        Heptane_Loss_Pct=excluded.Heptane_Loss_Pct,
                        Initial_LM_Pct=excluded.Initial_LM_Pct,
                        Initial_MA_Pct=excluded.Initial_MA_Pct,
                        Final_LM_Pct=excluded.Final_LM_Pct,
                        Final_MA_Pct=excluded.Final_MA_Pct,
                        SOP_Compliant=excluded.SOP_Compliant,
                        Deviation_Notes=excluded.Deviation_Notes,
                        Source_File=excluded.Source_File,
                        Last_Updated=CURRENT_TIMESTAMP
                ''', (
                    batch_no, process_type, start_dt, end_dt, final_input, final_output,
                    process_loss_pct, ma_conversion, heptane_loss, initial_lm_pct, initial_ma_pct, final_lm_pct, final_ma_pct, sop_compliant, "\n".join(notes), os.path.basename(filepath)
                ))
                
                # DB Insert Production_Log drums
                cursor.execute('DELETE FROM Production_Log WHERE Batch_Number = ?', (batch_no,))
                for d in input_drums:
                    cursor.execute('''
                        INSERT INTO Production_Log (Batch_Number, Stage, Material_Desc, Drum_Number, Drum_Weight, Production_Date)
                        VALUES (?, 'INPUT', ?, ?, ?, ?)
                    ''', (batch_no, d['desc'], d.get('drum_no', ''), d['weight'], start_dt))
                for d in output_drums:
                    cursor.execute('''
                        INSERT INTO Production_Log (Batch_Number, Stage, Material_Desc, Drum_Number, Drum_Weight, LM_GC, MA_GC, Heptane_GC, Production_Date)
                        VALUES (?, 'OUTPUT', ?, ?, ?, ?, ?, ?, ?)
                    ''', (batch_no, d['desc'], d.get('drum_no', ''), d['weight'], d['lm_gc'], d['ma_gc'], d['hpt_gc'], end_dt))
                
                i = j
            else:
                i += 1
            
    conn.commit()
    conn.close()
    print("Extraction and Upsert completed.")

if __name__ == "__main__":
    # Clear DB for fresh test
    conn = get_connection()
    conn.execute("DELETE FROM DPR_Master")
    conn.commit()
    conn.close()
    extract_and_upsert_dpr("/Users/abhi/WorkBench/Sharp_Mint/DPR/GLR_FY'26-27.xlsx")
