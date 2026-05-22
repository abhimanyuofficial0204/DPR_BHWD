import pandas as pd
import openpyxl
from datetime import datetime
import json
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def is_apr_may(dt):
    if not dt: return False
    if isinstance(dt, str):
        try: dt = pd.to_datetime(dt)
        except: return False
    if hasattr(dt, 'year'):
        return dt.year == 2026 and dt.month in [4, 5]
    return False

def get_dpr_raw(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sh in wb.sheetnames:
        rows = list(wb[sh].iter_rows(min_row=1, values_only=True))
        i = 0
        while i < len(rows):
            if rows[i] and rows[i][0] == 'Batch Size':
                start_dt = rows[i][8] if len(rows[i]) > 8 else None
                end_dt = rows[i+1][8] if i+1 < len(rows) and len(rows[i+1]) > 8 else None
                if is_apr_may(start_dt) or is_apr_may(end_dt):
                    # Found a batch! Collect rows until next 'Batch Size' or 'Description'
                    batch_rows = []
                    batch_id = rows[i][9] if len(rows[i]) > 9 else 'UNKNOWN'
                    proc_obj = rows[i][6] if len(rows[i]) > 6 else 'UNKNOWN'
                    
                    # Also grab the 2 rows before it to see the header
                    batch_rows.append(rows[max(0, i-2)])
                    batch_rows.append(rows[max(0, i-1)])
                    
                    j = i
                    while j < len(rows):
                        if j > i and rows[j] and rows[j][0] in ['Batch Size', 'Description', 'Drum no']:
                            break
                        batch_rows.append(rows[j])
                        j += 1
                        
                    return {'batch_id': batch_id, 'proc_obj': proc_obj, 'sheet': sh, 'rows': batch_rows}
            i += 1
    return None

def format_row(r):
    if not r: return ""
    return " | ".join([str(x).strip() if x is not None and str(x).strip()!="" else "" for x in r]).strip()

print("Extracting Sample 1: GLR")
glr_sample = get_dpr_raw("GLR_FY'26-27.xlsx")

print("Extracting Sample 2: Synthetic Reaction")
syn_sample = get_dpr_raw("Synthetic Reaction_FY'26-27.xlsx")

def find_feed_gc(batch_id, filepath):
    df = pd.read_excel(filepath, header=None)
    batch_rows = []
    capture = False
    
    # Try to match the last 3-4 digits
    import re
    m = re.search(r'[-/](\d+)$', str(batch_id))
    core_id = m.group(1) if m else str(batch_id)
    
    for idx, row in df.iterrows():
        col0 = str(row.iloc[0]).strip().upper()
        if pd.notna(row.iloc[0]) and ('#' in col0 or 'P-' in col0 or 'V-' in col0):
            if core_id in col0:
                capture = True
            else:
                if capture: break # end of block
        if capture:
            batch_rows.append(row.values.tolist())
    return batch_rows

print("Extracting matching Feed GC")
if glr_sample:
    glr_feed = find_feed_gc(glr_sample['batch_id'], "Feed GC Data/GLR-26-27.xlsx")
if syn_sample:
    syn_feed = find_feed_gc(syn_sample['batch_id'], "Feed GC Data/RXN-26-27.xlsx")

def find_prod(batch_id, filepath):
    df = pd.read_excel(filepath)
    import re
    m = re.search(r'[-/](\d+)$', str(batch_id))
    core_id = m.group(1) if m else str(batch_id)
    
    batch_rows = []
    batch_col = None
    for c in df.columns:
        if 'batch' in str(c).lower(): batch_col = c; break
    if not batch_col: return []
    
    # add header
    batch_rows.append(df.columns.tolist())
    
    for idx, row in df.iterrows():
        if pd.notna(row[batch_col]) and core_id in str(row[batch_col]):
            batch_rows.append(row.values.tolist())
    return batch_rows

print("Extracting matching Prod")
if glr_sample:
    glr_prod = find_prod(glr_sample['batch_id'], "Production Data/New Pro 26-27.xlsx")
if syn_sample:
    syn_prod = find_prod(syn_sample['batch_id'], "Production Data/New Pro 26-27.xlsx")

with open("raw_samples.txt", "w") as f:
    f.write("=== GLR SAMPLE ===\n")
    f.write(f"Batch ID: {glr_sample['batch_id']} | Process: {glr_sample['proc_obj']}\n")
    f.write("--- DPR ROWS ---\n")
    for r in glr_sample['rows']: f.write(format_row(r) + "\n")
    f.write("\n--- FEED GC ROWS ---\n")
    for r in glr_feed: f.write(format_row(r) + "\n")
    f.write("\n--- PROD ROWS ---\n")
    for r in glr_prod: f.write(format_row(r) + "\n")
    
    f.write("\n\n=== SYNTHETIC REACTION SAMPLE ===\n")
    f.write(f"Batch ID: {syn_sample['batch_id']} | Process: {syn_sample['proc_obj']}\n")
    f.write("--- DPR ROWS ---\n")
    for r in syn_sample['rows']: f.write(format_row(r) + "\n")
    f.write("\n--- FEED GC ROWS ---\n")
    for r in syn_feed: f.write(format_row(r) + "\n")
    f.write("\n--- PROD ROWS ---\n")
    for r in syn_prod: f.write(format_row(r) + "\n")

print("Done. Saved to raw_samples.txt")
