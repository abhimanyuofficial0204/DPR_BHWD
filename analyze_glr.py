import openpyxl
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

filepath = "/Users/abhi/WorkBench/Sharp_Mint/DPR/GLR_FY'26-27.xlsx"
wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

print("="*80)
print("1. SHEET NAMES")
print("="*80)
for i, name in enumerate(wb.sheetnames):
    print(f"  Sheet {i}: '{name}'")

print("\n" + "="*80)
print("2. HEADER ROW (Row 1) FOR EACH SHEET")
print("="*80)
for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if rows:
        print(f"\n--- Sheet: '{sh_name}' ---")
        for col_idx, val in enumerate(rows[0]):
            if val is not None and str(val).strip():
                print(f"  Col[{col_idx}] = {repr(val)}")

sh = wb.sheetnames[0]
all_rows = list(wb[sh].iter_rows(min_row=1, values_only=True))

print("\n" + "="*80)
print("3. FIRST BATCH IN APRIL 2026 - COMPLETE DUMP")
print("="*80)
from datetime import datetime
batch_positions = []
for i, r in enumerate(all_rows):
    if r and r[0] == 'Batch Size':
        batch_positions.append(i)

for bp in batch_positions:
    r = all_rows[bp]
    start_dt = r[8] if len(r) > 8 else None
    if start_dt and hasattr(start_dt, 'year') and start_dt.year == 2026 and start_dt.month == 4:
        print(f"\n4. SEPARATOR ROW (row {bp-1} before Batch Size):")
        if bp > 0:
            sep = all_rows[bp-1]
            for ci, v in enumerate(sep):
                if v is not None and str(v).strip() and str(v) != '0':
                    print(f"  Col[{ci}] = {repr(v)}")
            if all(v is None or str(v).strip() == '' or v == 0 for v in sep):
                print("  >> ALL CELLS ARE None/Empty/0 (black separator row)")
        
        print(f"\n3. BATCH DUMP starting at row {bp}:")
        end = batch_positions[batch_positions.index(bp) + 1] if batch_positions.index(bp) + 1 < len(batch_positions) else min(bp + 30, len(all_rows))
        for ri in range(bp, end):
            row = all_rows[ri]
            non_empty = [(ci, v) for ci, v in enumerate(row) if v is not None and str(v).strip() and str(v) != '0']
            if non_empty:
                print(f"  Row[{ri}]:")
                for ci, v in non_empty[:25]:
                    print(f"    Col[{ci}] = {repr(v)}")
            else:
                print(f"  Row[{ri}]: << all empty/zero >>")
        break

print("\n" + "="*80)
print("5. UNIQUE PROCESS OBJECTIVE VALUES (Col 6)")
print("="*80)
processes = set()
for sh_name in wb.sheetnames:
    rows = list(wb[sh_name].iter_rows(min_row=1, values_only=True))
    for r in rows:
        if r and r[0] == 'Batch Size' and len(r) > 6 and r[6] is not None:
            processes.add(str(r[6]).strip())
print(f"  Found {len(processes)} unique processes:")
for p in sorted(processes):
    print(f"    - '{p}'")

print("\n" + "="*80)
print("6. THREE BATCHES WITH DIFFERENT PROCESS OBJECTIVES")
print("="*80)
seen_procs = set()
count = 0
for sh_name in wb.sheetnames:
    if count >= 3:
        break
    rows = list(wb[sh_name].iter_rows(min_row=1, values_only=True))
    bp_list = [i for i, r in enumerate(rows) if r and r[0] == 'Batch Size']
    for idx, bp in enumerate(bp_list):
        if count >= 3:
            break
        r = rows[bp]
        proc = str(r[6]).strip() if len(r) > 6 and r[6] else 'UNKNOWN'
        if proc in seen_procs or proc == 'UNKNOWN':
            continue
        seen_procs.add(proc)
        count += 1
        
        end = bp_list[idx + 1] if idx + 1 < len(bp_list) else min(bp + 25, len(rows))
        print(f"\n--- Batch #{count}: Process='{proc}', Sheet='{sh_name}' ---")
        for ri in range(bp, min(bp + 15, end)):
            row = rows[ri]
            non_empty = [(ci, v) for ci, v in enumerate(row) if v is not None and str(v).strip() and str(v) != '0']
            if non_empty:
                print(f"  Row[{ri}]:")
                for ci, v in non_empty[:15]:
                    print(f"    Col[{ci}] = {repr(v)}")

wb.close()
print("\nDone!")
